"""Genera el cuadro comparativo de proveedores lado a lado."""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from . import estilos as E
from .modelo import Proveedor


def generar_comparativo_excel(titulo: str, proveedores: list[Proveedor], cfg: dict,
                              ruta: str, moneda: str = "S/.") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "COMPARATIVO"
    ws.sheet_view.showGridLines = False
    fmt = f'"{moneda}" #,##0.00'

    ws.column_dimensions["A"].width = 45
    n = max(1, len(proveedores))
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16

    # Titulo
    ws.merge_cells(f"A1:{get_column_letter(1 + n)}1")
    ws["A1"] = titulo or "CUADRO COMPARATIVO DE PROVEEDORES"
    ws["A1"].font = E.font(12, True, E.BLANCO); ws["A1"].fill = E.fill(E.AZUL)
    ws["A1"].alignment = E.center(True)
    ws.row_dimensions[1].height = 26

    # Cabecera de proveedores
    ws["A3"] = "DESCRIPCION / ITEM"; ws["A3"].font = E.font(9, True, E.BLANCO)
    ws["A3"].fill = E.fill(E.GRIS); ws["A3"].alignment = E.center()
    for i, p in enumerate(proveedores):
        col = get_column_letter(2 + i)
        ws[f"{col}3"] = p.nombre or f"Proveedor {i+1}"
        ws[f"{col}3"].font = E.font(9, True, E.BLANCO); ws[f"{col}3"].fill = E.fill(E.GRIS)
        ws[f"{col}3"].alignment = E.center(True)
    E.aplicar_borde(ws, f"A3:{get_column_letter(1+n)}3")

    # Union de descripciones (por texto) preservando orden
    items: list[str] = []
    for p in proveedores:
        for ln in p.lineas:
            d = ln.descripcion.strip()
            if d and d not in items:
                items.append(d)

    fila = 4
    for desc in items:
        ws[f"A{fila}"] = desc; ws[f"A{fila}"].font = E.font(8); ws[f"A{fila}"].alignment = E.left()
        valores = []
        for i, p in enumerate(proveedores):
            col = get_column_letter(2 + i)
            val = next((ln.costo_parcial() for ln in p.lineas if ln.descripcion.strip() == desc), None)
            cell = ws[f"{col}{fila}"]
            if val is not None:
                cell.value = val; cell.number_format = fmt; valores.append((val, col))
            else:
                cell.value = "No cotiza"; cell.font = E.font(8, color=E.GRIS)
            cell.alignment = E.right()
            cell.font = E.font(8)
        # resalta el menor de la fila
        if valores:
            _, col_min = min(valores, key=lambda t: t[0])
            ws[f"{col_min}{fila}"].fill = E.fill("C6EFCE")
        E.aplicar_borde(ws, f"A{fila}:{get_column_letter(1+n)}{fila}")
        ws.row_dimensions[fila].height = 26
        fila += 1

    # Fila TOTAL
    ws[f"A{fila}"] = "TOTAL"; ws[f"A{fila}"].font = E.font(9, True); ws[f"A{fila}"].fill = E.fill(E.AZUL_CLARO)
    totales = []
    for i, p in enumerate(proveedores):
        col = get_column_letter(2 + i)
        total = round(sum(ln.costo_parcial() for ln in p.lineas), 2)
        c = ws[f"{col}{fila}"]; c.value = total; c.number_format = fmt
        c.font = E.font(9, True); c.fill = E.fill(E.AZUL_CLARO); c.alignment = E.right()
        totales.append((total, col))
    if totales:
        _, col_min = min(totales, key=lambda t: t[0])
        ws[f"{col_min}{fila}"].fill = E.fill("C6EFCE")
    E.aplicar_borde(ws, f"A{fila}:{get_column_letter(1+n)}{fila}")

    # Nota
    fila += 2
    if totales:
        mejor = min(totales, key=lambda t: t[0])
        idx = ord(mejor[1]) - ord("B")
        ws[f"A{fila}"] = f"Proveedor mas economico: {proveedores[idx].nombre}  ->  {moneda} {mejor[0]:,.2f}"
        ws[f"A{fila}"].font = E.font(9, True, E.AZUL)

    ws.print_area = f"A1:{get_column_letter(1+n)}{fila}"
    ws.page_setup.orientation = "landscape"
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    wb.save(ruta)
    return ruta
