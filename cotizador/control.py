"""Registro de cotizaciones en la bitacora de la app y (opcional) en el maestro."""
from __future__ import annotations
import os
import datetime as dt
from openpyxl import Workbook, load_workbook

from . import estilos as E
from .modelo import Cotizacion

COLUMNAS = ["FECHA", "N° COTIZACION", "N° COT PROVEEDOR", "PROVEEDOR", "CLIENTE",
            "DESCRIPCION DEL SERVICIO", "VENTA s/IGV", "GASTO s/IGV",
            "UTILIDAD", "UTILIDAD %", "VENTA c/IGV"]


def _fila(cot: Cotizacion) -> list:
    prov = cot.proveedor_principal
    return [dt.date.today().isoformat(), cot.numero, prov.numero_cotizacion,
            prov.nombre, cot.cliente, cot.asunto, cot.subtotal(), cot.costo_gasto(),
            cot.utilidad(), round(cot.utilidad_pct(), 4), cot.total()]


def registrar_en_bitacora(cot: Cotizacion, ruta_bitacora: str) -> str:
    """Anexa una fila a la bitacora de la app (la crea si no existe)."""
    if os.path.exists(ruta_bitacora):
        wb = load_workbook(ruta_bitacora)
        ws = wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "BITACORA"
        ws.append(COLUMNAS)
        for i, _ in enumerate(COLUMNAS, 1):
            c = ws.cell(row=1, column=i)
            c.font = E.font(9, True, E.BLANCO); c.fill = E.fill(E.AZUL)
            c.alignment = E.center(True)
        widths = [12, 16, 16, 22, 20, 45, 13, 13, 12, 10, 13]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    fila = _fila(cot)
    ws.append(fila)
    r = ws.max_row
    for i in range(1, len(fila) + 1):
        c = ws.cell(row=r, column=i)
        if COLUMNAS[i - 1] in ("VENTA s/IGV", "GASTO s/IGV", "UTILIDAD", "VENTA c/IGV"):
            c.number_format = '"S/." #,##0.00'
        if COLUMNAS[i - 1] == "UTILIDAD %":
            c.number_format = '0.00%'
    os.makedirs(os.path.dirname(ruta_bitacora), exist_ok=True)
    wb.save(ruta_bitacora)
    return ruta_bitacora


def registrar_en_maestro(cot: Cotizacion, ruta_maestro: str, hoja: str,
                         mapeo: dict | None = None) -> str:
    """Anexa una fila a una hoja del Excel maestro existente.

    ADVERTENCIA: openpyxl puede no preservar tablas dinamicas/graficos del maestro.
    Se recomienda trabajar sobre una COPIA. 'mapeo' relaciona encabezado->valor.
    """
    wb = load_workbook(ruta_maestro)
    if hoja not in wb.sheetnames:
        raise ValueError(f"La hoja '{hoja}' no existe. Hojas: {wb.sheetnames}")
    ws = wb[hoja]

    # Detecta fila de encabezados (primera con >=3 celdas de texto)
    encabezados = {}
    fila_hdr = None
    for r in range(1, min(ws.max_row, 15) + 1):
        textos = [(c.column, str(c.value).strip()) for c in ws[r] if isinstance(c.value, str) and c.value.strip()]
        if len(textos) >= 3:
            fila_hdr = r
            encabezados = {v.upper(): col for col, v in textos}
            break

    valores = mapeo or {
        "N°  COTIZACION": cot.numero, "N° COTIZACION": cot.numero,
        "N° COT SEGUN PROVEE": cot.proveedor_principal.numero_cotizacion,
        "CLIENTE": cot.cliente,
        "DESCRIPCIÓN DEL SERVICIO": cot.asunto, "DESCRIPCION DEL SERVICIO": cot.asunto,
        "PROVEEDOR": cot.proveedor_principal.nombre,
    }
    destino = ws.max_row + 1
    for k, v in valores.items():
        col = encabezados.get(k.upper())
        if col:
            ws.cell(row=destino, column=col, value=v)
    wb.save(ruta_maestro)
    return ruta_maestro
