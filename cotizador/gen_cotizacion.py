"""Genera la Cotizacion FMI en Excel (formato PAN-FO-FM-02)."""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from . import estilos as E
from .config import ruta_logo
from .modelo import Cotizacion


def _mon(v, moneda="S/."):
    return f'"{moneda}" #,##0.00'


def generar_cotizacion_excel(cot: Cotizacion, cfg: dict, ruta: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "COTIZACION"
    ws.sheet_view.showGridLines = False

    emp = cfg["empresa"]
    cond = cfg["condiciones_comerciales"]
    notas = cfg.get("notas", [])
    fmt = _mon(0, cot.moneda)

    # Anchos de columna A..Q
    anchos = {"A": 6, "B": 10, "C": 10, "D": 10, "E": 10, "F": 8, "G": 8, "H": 6,
              "I": 6, "J": 7, "K": 3, "L": 11, "M": 13, "N": 5, "O": 3, "P": 13, "Q": 13}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # ---------- Encabezado ----------
    ws.merge_cells("A1:C4")
    ws["A1"].alignment = E.center(True)
    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 16
    logo = ruta_logo(cfg)
    if not E.insertar_logo(ws, logo, "A2", max_w=175, max_h=52):
        ws["A1"] = "PANORAMA\nOutsourcing"
        ws["A1"].font = E.font(14, True, E.NARANJA)

    ws.merge_cells("D1:M4")
    ws["D1"] = "COTIZACION DE SERVICIOS"
    ws["D1"].font = E.font(14, True)
    ws["D1"].alignment = E.center()

    ws.merge_cells("N1:O1"); ws["N1"] = "Codigo:"; ws["P1"] = emp["codigo_cotizacion"]
    ws.merge_cells("N2:O2"); ws["N2"] = "Version:"; ws["P2"] = emp["version_cotizacion"]
    ws.merge_cells("N3:O3"); ws["N3"] = "Pagina:"; ws["P3"] = "1 de 1"
    for r in (1, 2, 3):
        ws.merge_cells(f"P{r}:Q{r}")
        ws[f"N{r}"].font = E.font(8, True); ws[f"N{r}"].alignment = E.center()
        ws[f"P{r}"].font = E.font(8); ws[f"P{r}"].alignment = E.center()
    E.aplicar_borde(ws, "N1:Q3")
    E.aplicar_borde(ws, "A1:Q4", E.BORDE)

    # ---------- Datos generales ----------
    datos = [
        ("Fecha:", cot.fecha),
        ("Cliente:", cot.cliente),
        ("Contacto:", cot.contacto),
        ("Asunto :", cot.asunto),
    ]
    fila = 6
    for etq, val in datos:
        ws[f"A{fila}"] = etq; ws[f"A{fila}"].font = E.font(9, True)
        ws.merge_cells(f"C{fila}:Q{fila}")
        ws[f"C{fila}"] = val; ws[f"C{fila}"].font = E.font(9); ws[f"C{fila}"].alignment = E.left()
        fila += 1

    # Referencia / N° presupuesto
    fila += 0
    ws.merge_cells(f"A{fila}:H{fila}")
    ws[f"A{fila}"] = "Referencia"; ws[f"A{fila}"].font = E.font(9, True, E.BLANCO)
    ws[f"A{fila}"].fill = E.fill(E.GRIS); ws[f"A{fila}"].alignment = E.left(False)
    ws.merge_cells(f"I{fila}:N{fila}")
    ws[f"I{fila}"] = "Presupuesto Economico N°"; ws[f"I{fila}"].font = E.font(9, True, E.BLANCO)
    ws[f"I{fila}"].fill = E.fill(E.GRIS); ws[f"I{fila}"].alignment = E.center()
    E.aplicar_borde(ws, f"A{fila}:Q{fila}")
    fila += 1
    ws.merge_cells(f"A{fila}:B{fila}")
    ws[f"A{fila}"] = "N° Ticket / Solicitud"; ws[f"A{fila}"].font = E.font(8, True)
    ws.merge_cells(f"C{fila}:H{fila}")
    ws[f"C{fila}"] = cot.referencia_ticket; ws[f"C{fila}"].font = E.font(9); ws[f"C{fila}"].alignment = E.left()
    ws.merge_cells(f"I{fila}:N{fila}")
    ws[f"I{fila}"] = cot.numero; ws[f"I{fila}"].font = E.font(9, True); ws[f"I{fila}"].alignment = E.center()
    E.aplicar_borde(ws, f"A{fila}:Q{fila}")

    # ---------- Cabecera de tabla ----------
    fila += 2
    hdr = fila
    cabeceras = [("A", "ITEM"), ("B", "D E S C R I P C I O N"), ("I", "U.M."),
                 ("J", "CANT."), ("L", "PRECIO"), ("M", "Precio Parcial"),
                 ("P", "PRECIO SUB\nTOTAL")]
    ws.merge_cells(f"B{hdr}:H{hdr}")
    ws.merge_cells(f"L{hdr}:L{hdr}")
    ws.merge_cells(f"M{hdr}:O{hdr}")
    ws.merge_cells(f"P{hdr}:Q{hdr}")
    for col, txt in cabeceras:
        c = ws[f"{col}{hdr}"]
        c.value = txt; c.font = E.font(8, True, E.BLANCO)
        c.fill = E.fill(E.AZUL); c.alignment = E.center(True)
    E.aplicar_borde(ws, f"A{hdr}:Q{hdr}")
    ws.row_dimensions[hdr].height = 26

    # ---------- Grupo 01.00 (titulo) ----------
    fila += 1
    grupo = fila
    ws[f"A{grupo}"] = "01.00"; ws[f"A{grupo}"].font = E.font(8, True, color=E.AZUL)
    ws.merge_cells(f"B{grupo}:O{grupo}")
    ws[f"B{grupo}"] = (cot.asunto or "").upper()
    ws[f"B{grupo}"].font = E.font(8, True, E.AZUL); ws[f"B{grupo}"].alignment = E.left()
    ws[f"B{grupo}"].fill = E.fill(E.AZUL_CLARO)
    ws.merge_cells(f"P{grupo}:Q{grupo}")
    ws[f"P{grupo}"].number_format = fmt; ws[f"P{grupo}"].font = E.font(9, True)
    ws[f"P{grupo}"].alignment = E.right()
    E.aplicar_borde(ws, f"A{grupo}:Q{grupo}")

    # ---------- Lineas ----------
    fila += 1
    primera = fila
    for i, ln in enumerate(cot.lineas, start=1):
        vu = ln.venta_unit(cot.margen, cot.modo_margen)
        vp = ln.venta_parcial(cot.margen, cot.modo_margen)
        ws[f"A{fila}"] = f"01.{i:02d}"; ws[f"A{fila}"].font = E.font(8); ws[f"A{fila}"].alignment = E.center()
        ws.merge_cells(f"B{fila}:H{fila}")
        ws[f"B{fila}"] = ln.descripcion; ws[f"B{fila}"].font = E.font(8); ws[f"B{fila}"].alignment = E.left()
        ws[f"I{fila}"] = ln.unidad; ws[f"I{fila}"].font = E.font(8); ws[f"I{fila}"].alignment = E.center()
        ws[f"J{fila}"] = ln.cantidad; ws[f"J{fila}"].font = E.font(8); ws[f"J{fila}"].alignment = E.center()
        ws[f"L{fila}"] = vu; ws[f"L{fila}"].number_format = fmt; ws[f"L{fila}"].font = E.font(8); ws[f"L{fila}"].alignment = E.right()
        ws.merge_cells(f"M{fila}:O{fila}")
        ws[f"M{fila}"] = vp; ws[f"M{fila}"].number_format = fmt; ws[f"M{fila}"].font = E.font(8); ws[f"M{fila}"].alignment = E.right()
        ws.merge_cells(f"P{fila}:Q{fila}")
        ws[f"P{fila}"] = vp; ws[f"P{fila}"].number_format = fmt; ws[f"P{fila}"].font = E.font(8); ws[f"P{fila}"].alignment = E.right()
        ws.row_dimensions[fila].height = 30
        E.aplicar_borde(ws, f"A{fila}:Q{fila}")
        fila += 1
    ultima = fila - 1

    # subtotal del grupo 01.00
    ws[f"P{grupo}"] = f"=SUM(P{primera}:P{ultima})"

    # ---------- Totales ----------
    def fila_total(etq, valor, resaltar=False):
        nonlocal fila
        ws.merge_cells(f"L{fila}:O{fila}")
        c = ws[f"L{fila}"]; c.value = etq
        c.font = E.font(9, True, E.BLANCO if resaltar else "000000")
        c.alignment = E.left(False)
        if resaltar:
            c.fill = E.fill(E.AZUL)
            for col in "LMNOPQ":
                ws[f"{col}{fila}"].fill = E.fill(E.AZUL)
        ws.merge_cells(f"P{fila}:Q{fila}")
        ws[f"P{fila}"] = valor; ws[f"P{fila}"].number_format = fmt
        ws[f"P{fila}"].font = E.font(9, True, E.BLANCO if resaltar else "000000")
        ws[f"P{fila}"].alignment = E.right()
        E.aplicar_borde(ws, f"L{fila}:Q{fila}")
        fila += 1

    fila_total("COSTO DIRECTO", f"=SUM(P{primera}:P{ultima})")
    fila_total(f"IGV {int(cot.igv*100)}%", f"=P{fila-1}*{cot.igv}")
    fila_total("PRECIO CON IGV", f"=P{fila-2}+P{fila-1}", resaltar=True)

    # ---------- Condiciones comerciales ----------
    fila += 1
    ws[f"A{fila}"] = "Condiciones Comerciales:"; ws[f"A{fila}"].font = E.font(9, True)
    fila += 1
    condiciones = [
        ("Validez de la Propuesta Economica:", cond["validez"]),
        ("Forma de pago:", cond["forma_pago"]),
        ("Plazo de Ejecucion y programacion:", cond["plazo_ejecucion"]),
        ("Tiempo de ejecucion:", cond["tiempo_ejecucion"]),
        ("Lugar de Entrega:", cond["lugar_entrega"]),
        ("Orden a :", cond["orden_a"]),
    ]
    for i, (k, v) in enumerate(condiciones, 1):
        ws[f"A{fila}"] = f"0{i}.00"; ws[f"A{fila}"].font = E.font(8, True)
        ws.merge_cells(f"B{fila}:F{fila}")
        ws[f"B{fila}"] = k; ws[f"B{fila}"].font = E.font(8, True)
        ws.merge_cells(f"G{fila}:Q{fila}")
        ws[f"G{fila}"] = v; ws[f"G{fila}"].font = E.font(8); ws[f"G{fila}"].alignment = E.left()
        fila += 1

    # ---------- Notas ----------
    fila += 1
    ws[f"A{fila}"] = "Notas:"; ws[f"A{fila}"].font = E.font(9, True)
    fila += 1
    for i, nota in enumerate(notas, 1):
        ws[f"A{fila}"] = f"0{i}.00"; ws[f"A{fila}"].font = E.font(8, True)
        ws.merge_cells(f"B{fila}:Q{fila}")
        ws[f"B{fila}"] = nota; ws[f"B{fila}"].font = E.font(8); ws[f"B{fila}"].alignment = E.left()
        ws.row_dimensions[fila].height = 26
        fila += 1

    # ---------- Firma ----------
    fila += 3
    ws.merge_cells(f"L{fila}:Q{fila}")
    from .config import ruta_firma
    E.insertar_logo(ws, ruta_firma(cfg), f"L{fila-2}", max_w=150, max_h=48)
    ws[f"L{fila}"] = "_______________________________"; ws[f"L{fila}"].alignment = E.center()
    fila += 1
    ws.merge_cells(f"L{fila}:Q{fila}")
    ws[f"L{fila}"] = emp["firma_nombre"]; ws[f"L{fila}"].font = E.font(9, True); ws[f"L{fila}"].alignment = E.center()
    fila += 1
    ws.merge_cells(f"L{fila}:Q{fila}")
    ws[f"L{fila}"] = emp["firma_cargo"]; ws[f"L{fila}"].font = E.font(8); ws[f"L{fila}"].alignment = E.center()

    ws.print_area = f"A1:Q{fila}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    wb.save(ruta)
    return ruta
