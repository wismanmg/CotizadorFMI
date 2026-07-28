"""Helpers de estilo para openpyxl (colores y bordes de la marca Panorama)."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AZUL = "1F3864"
AZUL_CLARO = "D9E1F2"
GRIS = "808080"
GRIS_CLARO = "F2F2F2"
NARANJA = "C55A11"
BLANCO = "FFFFFF"

FONT = "Calibri"

def font(size=9, bold=False, color="000000"):
    return Font(name=FONT, size=size, bold=bold, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right(wrap=False):
    return Alignment(horizontal="right", vertical="center", wrap_text=wrap)

_thin = Side(style="thin", color="BFBFBF")
_med = Side(style="medium", color=AZUL)
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDE_GRUESO = Border(left=_med, right=_med, top=_med, bottom=_med)

def aplicar_borde(ws, rango, borde=BORDE):
    for fila in ws[rango]:
        for c in fila:
            c.border = borde


def insertar_logo(ws, ruta, ancla, max_w=170, max_h=64):
    """Inserta un logo escalado (manteniendo proporcion) en la celda 'ancla'.
    Devuelve True si lo inserto. No falla si la imagen no existe."""
    if not ruta:
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        with PILImage.open(ruta) as im:
            w, h = im.size
        escala = min(max_w / w, max_h / h, 1.0)
        img = XLImage(ruta)
        img.width = int(w * escala)
        img.height = int(h * escala)
        ws.add_image(img, ancla)
        return True
    except Exception:
        return False
