"""Genera el Sustento de Ingresos y Gastos (formato PAN-FO-FM-11)."""
from __future__ import annotations
import os
from openpyxl import Workbook

from . import estilos as E
from .config import ruta_logo
from .modelo import Cotizacion


def generar_sustento_excel(cot: Cotizacion, cfg: dict, ruta: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "SUSTENTO"
    ws.sheet_view.showGridLines = False
    emp = cfg["empresa"]
    fmt = f'"{cot.moneda}" #,##0.00'
    pct = '0.00%'

    for col, w in {"A": 3, "B": 12, "C": 14, "D": 12, "E": 12, "F": 4,
                   "G": 12, "H": 10, "I": 10, "J": 12, "K": 6}.items():
        ws.column_dimensions[col].width = w

    # Encabezado (logo | titulo | codigo/version/pagina) dentro de un recuadro
    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 18
    ws.merge_cells("A1:B3")   # caja del logo
    ws["A1"].alignment = E.center(True)
    logo = ruta_logo(cfg)
    if not E.insertar_logo(ws, logo, "A1", max_w=150, max_h=52):
        ws["A1"] = "PANORAMA\nOutsourcing"
        ws["A1"].font = E.font(12, True, E.NARANJA)

    ws.merge_cells("C1:G3")
    ws["C1"] = "SUSTENTO DE INGRESOS Y GASTOS"
    ws["C1"].font = E.font(14, True); ws["C1"].alignment = E.center(True)

    caja = [("Codigo:", emp["codigo_sustento"]), ("Version:", emp["version_sustento"]),
            ("Pagina:", "1 de 1")]
    for i, (k, v) in enumerate(caja, start=1):
        ws[f"H{i}"] = k; ws[f"H{i}"].font = E.font(8, True); ws[f"H{i}"].alignment = E.right()
        ws.merge_cells(f"I{i}:J{i}")
        ws[f"I{i}"] = v; ws[f"I{i}"].font = E.font(8); ws[f"I{i}"].alignment = E.center()
    E.aplicar_borde(ws, "A1:J3", E.BORDE)

    ws["A5"] = "Fecha:"; ws["A5"].font = E.font(10, True)
    ws["E5"] = cot.fecha; ws["E5"].font = E.font(10)
    ws["A7"] = "Presupuesto Economico de referencia"; ws["A7"].font = E.font(10, True)
    ws["E7"] = cot.numero; ws["E7"].font = E.font(10, True)

    def titulo(fila, txt, color):
        ws.merge_cells(f"C{fila}:J{fila}")
        c = ws[f"C{fila}"]; c.value = txt
        c.font = E.font(11, True, E.BLANCO); c.fill = E.fill(color)
        c.alignment = E.center()

    def dato(fila, etq, val, bold=False):
        ws[f"C{fila}"] = etq; ws[f"C{fila}"].font = E.font(10, True)
        ws.merge_cells(f"G{fila}:J{fila}")
        ws[f"G{fila}"] = val
        ws[f"G{fila}"].font = E.font(10, bold); ws[f"G{fila}"].alignment = E.left(False)

    def monto(fila, etq, val, resaltar=False):
        ws[f"D{fila}"] = etq; ws[f"D{fila}"].font = E.font(10, resaltar)
        ws[f"G{fila}"] = val; ws[f"G{fila}"].number_format = fmt
        ws[f"G{fila}"].font = E.font(10, resaltar); ws[f"G{fila}"].alignment = E.right()

    # ---- GASTO ----
    prov = cot.proveedor_principal
    titulo(10, "GASTO", E.NARANJA)
    dato(12, "Mes :", cot.fecha.split(",")[-1].strip() if "," in cot.fecha else "")
    dato(13, "MOTIVO :", cot.asunto)
    dato(15, "Proveedor :", prov.nombre or "(proveedor)")
    gasto = cot.costo_gasto()
    monto(17, "Monto", gasto)
    monto(18, "IGV", round(gasto * cot.igv, 2))
    monto(19, "Total c/IGV", cot.gasto_con_igv(), resaltar=True)

    # ---- VENTA ----
    titulo(22, "VENTA", E.AZUL)
    dato(24, "Cliente :", cot.cliente)
    dato(25, "Local / Sede :", cot.sede or cot.asunto)
    venta = cot.subtotal()
    monto(27, "Monto", venta)
    monto(28, "IGV", cot.igv_monto())
    monto(29, "Monto c/IGV", cot.total(), resaltar=True)

    # ---- RESUMEN ----
    titulo(32, "RESUMEN", E.GRIS)
    ws["G34"] = "s/IGV"; ws["G34"].font = E.font(9, True); ws["G34"].alignment = E.center()
    ws["J34"] = "c/IGV"; ws["J34"].font = E.font(9, True); ws["J34"].alignment = E.center()

    def resumen(fila, etq, sin_igv, con_igv, fmt_col=fmt):
        ws[f"C{fila}"] = etq; ws[f"C{fila}"].font = E.font(10, True)
        ws[f"G{fila}"] = sin_igv; ws[f"G{fila}"].number_format = fmt_col; ws[f"G{fila}"].alignment = E.right()
        if con_igv is not None:
            ws[f"J{fila}"] = con_igv; ws[f"J{fila}"].number_format = fmt_col; ws[f"J{fila}"].alignment = E.right()

    resumen(35, "Total Venta", venta, cot.total())
    resumen(36, "Total Gasto", gasto, cot.gasto_con_igv())
    resumen(37, "Utilidad Obtenida", cot.utilidad(), round(cot.utilidad() * (1 + cot.igv), 2))
    ws["C39"] = "Utilidad Obtenida (%)"; ws["C39"].font = E.font(10, True)
    ws["G39"] = cot.utilidad_pct(); ws["G39"].number_format = pct; ws["G39"].alignment = E.right()

    # Firma
    from .config import ruta_firma
    E.insertar_logo(ws, ruta_firma(cfg), "G44", max_w=150, max_h=48)
    ws.merge_cells("G47:J47"); ws["G47"] = "_____________________________"; ws["G47"].alignment = E.center()
    ws.merge_cells("G48:J48"); ws["G48"] = emp["firma_nombre"]; ws["G48"].font = E.font(9, True); ws["G48"].alignment = E.center()
    ws.merge_cells("G49:J49"); ws["G49"] = emp["firma_cargo"]; ws["G49"].font = E.font(8); ws["G49"].alignment = E.center()

    ws.print_area = "A1:K50"
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    wb.save(ruta)
    return ruta
